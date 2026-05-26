"""
Remplissage professionnel du modele Analyse Financiere Pro
Projet: EpitopX - Plateforme SaaS Bioinformatique
Startup Algerienne | 2025-2026
VERSION REALISTE - Couts fixes <= 3 000 000 DA/an
"""
import openpyxl
import shutil

SRC  = r'C:\Users\asus\Downloads\Analyse financiere pro (1).xlsx'
DST  = r'C:\Users\asus\Downloads\EpitopX_Analyse_Financiere_v8.xlsx'

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

# helpers
def sv(ws, row, col, val):
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    if isinstance(cell.value, str) and cell.value.startswith('='):
        return
    cell.value = val

def force(ws, row, col, val):
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = val

# ============================================================================
# 1. CHARGES VARIABLES  (cout par abonnement - lot de reference: 1 000 abos)
# ============================================================================
ws_cv = wb['Charges variables']

infra = [
    # Costs per 1000 subs/month — paid tiers kick in above ~100 users
    (5,  "Appels API LLM (OpenRouter - paid tier from M6)",   1000,"abonnements", 22400, 1, "appel/abo"),
    (6,  "Compute cloud (Render Starter $25/mois from M5)",   1000,"abonnements",  8400, 1, "instance"),
    (7,  "Stockage donnees utilisateurs (Cloudflare R2)",     1000,"abonnements",  2800, 1, "GB/abo"),
    (8,  "CDN & bande passante (Cloudflare free)",            1000,"abonnements",     0, 1, "-"),
    (9,  "Emails transactionnels (Brevo - plan gratuit)",     1000,"abonnements",     0, 1, "-"),
    (10, "Frais passerelle paiement (Stripe 2.9%+0.30$)",     1000,"abonnements", 73080, 1, "trans./abo"),
    (11, "Monitoring erreurs (Sentry - free tier)",           1000,"abonnements",     0, 1, "-"),
    (12, "Analytics produit (PostHog - free tier)",           1000,"abonnements",     0, 1, "-"),
    (13, "-", 1, None, 0, 0, None), (14, "-", 1, None, 0, 0, None),
    (15, "-", 1, None, 0, 0, None), (16, "-", 1, None, 0, 0, None),
    (17, "-", 1, None, 0, 0, None), (18, "-", 1, None, 0, 0, None),
    (19, "-", 1, None, 0, 0, None), (20, "-", 1, None, 0, 0, None),
    (21, "-", 1, None, 0, 0, None), (22, "-", 1, None, 0, 0, None),
    (23, "-", 1, None, 0, 0, None), (24, "-", 1, None, 0, 0, None),
    (25, "-", 1, None, 0, 0, None), (26, "-", 1, None, 0, 0, None),
]
consomm = [
    (27, "Support client (Crisp - plan gratuit)",     1000,"abonnements", 0, 1, "-"),
    (28, "Backup automatise (GitHub + Google Drive)",  1000,"abonnements", 0, 1, "-"),
    (29, "CI/CD (GitHub Actions - free tier)",         1000,"abonnements", 0, 1, "-"),
    (30, "-", 1, None, 0, 0, None), (31, "-", 1, None, 0, 0, None),
    (32, "-", 1, None, 0, 0, None), (33, "-", 1, None, 0, 0, None),
    (34, "-", 1, None, 0, 0, None), (35, "-", 1, None, 0, 0, None),
    (36, "-", 1, None, 0, 0, None), (37, "-", 1, None, 0, 0, None),
    (38, "-", 1, None, 0, 0, None), (39, "-", 1, None, 0, 0, None),
    (40, "-", 1, None, 0, 0, None), (41, "-", 1, None, 0, 0, None),
    (42, "-", 1, None, 0, 0, None), (43, "-", 1, None, 0, 0, None),
    (44, "-", 1, None, 0, 0, None), (45, "-", 1, None, 0, 0, None),
    (46, "-", 1, None, 0, 0, None), (47, "-", 1, None, 0, 0, None),
]
autres = [(r, "-", 1, None, 0, 0, None) for r in range(48, 55)]

for items in (infra, consomm, autres):
    for r, nom, qty, unite, total, q_unit, u_unit in items:
        sv(ws_cv, r, 4, nom);   sv(ws_cv, r, 5, qty)
        sv(ws_cv, r, 6, unite); sv(ws_cv, r, 7, total)
        sv(ws_cv, r, 8, q_unit); sv(ws_cv, r, 9, u_unit)

force(ws_cv, 4,  3, "Infrastructure Cloud & API")
force(ws_cv, 26, 3, "Services numeriques")
force(ws_cv, 47, 3, "Autres")
print("OK Charges variables rempli")

# ============================================================================
# 2. CHARGES FIXES - Total annuel <= 3 000 000 DA
# ============================================================================
ws_cf = wb['Charges fixes']

equipe = [
    (4,  "Chef de Projet & Lead Developpeur Full-Stack", 1, 60000),
    (5,  "Developpeur Backend Python / Django",           1, 60000),
    (6,  "-",                                             0,     0),
    (7,  "-",                                             0,     0),
    (8,  "-",                                             0,     0),
    (9,  "Specialiste Theileria / Parasitologue",         1, 60000),
    (10, "Responsable Marketing & Ventes",                1, 60000),
    (11, "-", 0, 0), (12, "-", 0, 0), (13, "-", 0, 0),
    (14, "-", 0, 0), (15, "-", 0, 0), (16, "-", 0, 0),
    (17, "-", 0, 0), (18, "-", 0, 0), (19, "-", 0, 0),
    (20, "-", 0, 0), (21, "-", 0, 0), (22, "-", 0, 0), (23, "-", 0, 0),
]
for r, poste, nb, sal in equipe:
    sv(ws_cf, r, 3, poste); sv(ws_cf, r, 4, nb); sv(ws_cf, r, 5, sal)

services = [
    # Render paid plan avg ~$18/mois (free first 4 months, $25 from M5)
    (30, "Hebergement Render (avg $18/mois - passage paye M5)", 1, 2520),
    # LLM paid avg ~$15/mois (free first 5 months, paid from M6)
    (31, "API LLM OpenRouter (avg $15/mois - paye from M6)",    1, 2100),
    (32, "Nom de domaine .com (Namecheap)",                     1,  980),
    (33, "GitHub Free / outils open-source",                    1,    0),
    (34, "Travail 100% a distance (0 loyer)",                   1,    0),
    (35, "Internet (forfait personnel equipe)",                 1,    0),
    (36, "Comptabilite (Wave - logiciel gratuit)",              1,    0),
    (37, "Assurance (souscription annee 2)",                    1,    0),
    (38, "-", 0, 0), (39, "-", 0, 0), (40, "-", 0, 0),
    (41, "-", 0, 0), (42, "-", 0, 0), (43, "-", 0, 0),
    (44, "-", 0, 0), (45, "-", 0, 0), (46, "-", 0, 0),
    (47, "-", 0, 0), (48, "-", 0, 0), (49, "-", 0, 0),
]
for r, nom, nb, mnt in services:
    sv(ws_cf, r, 3, nom); sv(ws_cf, r, 4, nb); sv(ws_cf, r, 5, mnt)

print("OK Charges fixes rempli  [annuel estime: 2 947 200 DA < 3 000 000 DA]")

# ============================================================================
# 3. EQUIPEMENT & INVESTISSEMENT
# ============================================================================
ws_eq = wb['Equipement & investissement']

equipements = [
    (4,  "PC Portables developpement (i5/16GB, reconditionnes)", 60000, 3),
    (5,  "Disques SSD externes - sauvegarde",                     6000, 3),
    (6,  "Casques + webcams (reunions distantes)",                 3000, 3),
    (7,  "-", 0, 0), (8,  "-", 0, 0), (9,  "-", 0, 0),
    (10, "-", 0, 0), (11, "-", 0, 0), (12, "-", 0, 0),
    (13, "-", 0, 0), (14, "-", 0, 0), (15, "-", 0, 0),
    (16, "-", 0, 0), (17, "-", 0, 0), (18, "-", 0, 0),
    (19, "-", 0, 0), (20, "-", 0, 0), (21, "-", 0, 0),
    (22, "-", 0, 0), (23, "-", 0, 0),
]
for r, nom, val_u, qty in equipements:
    sv(ws_eq, r, 3, nom); sv(ws_eq, r, 4, val_u); sv(ws_eq, r, 5, qty)

agencement = [
    (4,  "Configuration infrastructure cloud (par l'equipe)",  2000),
    (5,  "Identite visuelle - Canva + Figma (free)",               0),
    (6,  "Site marketing - GitHub Pages (gratuit)",                0),
    (7,  "Enregistrement juridique startup Algerie",           15000),
    (8,  "Outils dev - VS Code / Git / Docker (open-source)",      0),
    (9,  "Formation - docs officielles / YouTube (gratuit)",       0),
    (10, "Securite - Cloudflare + Let's Encrypt (gratuit)",        0),
    (11, "Campagne lancement - reseaux sociaux organiques",        0),
    (12, "Divers & imprevus (bugs, pannes, legaux)",          25000),
    (13, "-", 0), (14, "-", 0), (15, "-", 0),
    (16, "-", 0), (17, "-", 0), (18, "-", 0),
    (19, "-", 0), (20, "-", 0), (21, "-", 0),
    (22, "-", 0), (23, "-", 0),
]
for r, nom, val in agencement:
    sv(ws_eq, r, 10, nom); sv(ws_eq, r, 12, val)

print("OK Equipements rempli  [investissement total: 222 000 DA]")

# ============================================================================
# 4. MARKETING
# ============================================================================
ws_mk = wb['Marketing']
marketing_monthly = [50, 40, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35]
for i, budget in enumerate(marketing_monthly):
    sv(ws_mk, 5, 6 + i, budget)
print("OK Marketing rempli  [annuel: 440 000 DA (~$250/mois)]")

# ============================================================================
# 5. CHIFFRE D'AFFAIRES
# ============================================================================
ws_ca = wb["Chiffre d'affaires"]

sv(ws_ca, 5, 2, "EpitopX - Abonnements SaaS (Basic / Pro / Pro Plus)")
sv(ws_ca, 5, 3, "Abonnements vendus / mois")
sv(ws_ca, 5, 4, "Q1")
sv(ws_ca, 7, 3, "Prix moyen pondere (x1000 DA)")
ws_ca.cell(row=7, column=4).value = "P1"

prix_moyen = 2.52
for col in range(5, 17):
    sv(ws_ca, 7, col, prix_moyen)

# Conservative realistic growth for niche scientific SaaS - year 1
# M1: beta lab/friends  M2-M4: early adopters  M5-M12: steady ~25-30/month organic
abonnements = [5, 15, 30, 55, 85, 120, 160, 200, 245, 290, 345, 400]
for i, val in enumerate(abonnements):
    sv(ws_ca, 5, 5 + i, val)

print("OK Chiffre d'affaires rempli")

# ============================================================================
# 6. BILAN FINAL
# ============================================================================
ws_bf = wb['Bilan Final']
ws_bf.cell(row=1, column=2).value = "EpitopX - Plateforme Bioinformatique SaaS"
ws_bf.cell(row=2, column=2).value = "Analyse Financiere Previsionnelle - Annee 1 | 2025-2026"
print("OK Bilan Final annote")

# ============================================================================
# SAVE + RESUME
# ============================================================================
wb.save(DST)
print(f"\nFichier sauvegarde: {DST}")

total_subs   = sum(abonnements)
ca_annuel_DA = sum(s * prix_moyen * 1000 for s in abonnements)
charges_fixes = (4 * 60000 + 2520 + 2100 + 980) * 12   # 2,947,200 DA
var_per_sub  = (22400 + 8400 + 2800 + 73080) / 1000    # 106.68 DA/sub
var_total    = total_subs * var_per_sub
marketing_DA = sum(marketing_monthly) * 1000
equip_total  = 3*60000 + 3*6000 + 3*3000 + 2000 + 15000 + 25000  # 249,000 DA
total_couts  = charges_fixes + var_total + marketing_DA + equip_total
benefice     = ca_annuel_DA - total_couts

print("\n" + "="*60)
print("  RESUME FINANCIER EPITOPX - ANNEE 1")
print("="*60)
print(f"  Abonnements total (12 mois) : {total_subs:,}")
print(f"  Pic mensuel (mois 12)       : {abonnements[-1]:,} abonnes")
print(f"  Prix moyen                  : 2 520 DA ($18)")
print(f"  CA annuel                   : {ca_annuel_DA:>15,.0f} DA")
print(f"                                ${ca_annuel_DA/140:>13,.0f}")
print(f"  --------------------------------------------------")
print(f"  Salaires (4x60k x12)        : {4*60000*12:>15,.0f} DA")
print(f"  Services (Render+LLM+domaine): {(2520+2100+980)*12:>14,.0f} DA")
print(f"  Total charges fixes/an      : {charges_fixes:>15,.0f} DA  (< 3 000 000)")
print(f"  Cout variable/abonnement    : {var_per_sub:>15.0f} DA")
print(f"  Couts variables total       : {var_total:>15,.0f} DA")
print(f"  Marketing annuel            : {marketing_DA:>15,.0f} DA")
print(f"  Investissement equipement   : {equip_total:>15,.0f} DA")
print(f"  --------------------------------------------------")
print(f"  TOTAL COUTS ANNUELS         : {total_couts:>15,.0f} DA")
print(f"  BENEFICE ANNUEL             : {benefice:>15,.0f} DA")
print(f"                                ${benefice/140:>13,.0f}")
print(f"  MARGE BENEFICIAIRE          : {benefice/ca_annuel_DA*100:>14.1f}%")
print("="*60)
