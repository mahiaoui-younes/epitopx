import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\asus\Downloads\Analyse financiere pro (1).xlsx')

ws = wb["Chiffre d'affaires"]
print("=== CA Sheet (all non-empty rows with row number) ===")
for row_num in range(1, ws.max_row + 1):
    row_vals = [ws.cell(row=row_num, column=col).value for col in range(1, 20)]
    if any(v is not None for v in row_vals):
        print(f"Row {row_num}: {row_vals}")
