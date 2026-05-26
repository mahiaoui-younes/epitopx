import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\asus\Downloads\Analyse financiere pro (1).xlsx')

for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"\n=== {sh} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(c is not None for c in row):
            print(row)
